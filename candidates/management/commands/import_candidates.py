"""Import examinees from the exam-centre spreadsheet.

    python manage.py import_candidates data/islamabad.xlsx

Every row is upserted on `roll_number`, so re-running is safe. Each candidate
also gets a generated placeholder photo in `image` (base64 PNG) — sample data
that stands in until real photographs are supplied.
"""

from __future__ import annotations

import base64
import datetime as dt
import io
from pathlib import Path
from typing import Any

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont

from candidates.models import Candidate

# Spreadsheet header -> model field. Headers are normalised (lower-cased,
# whitespace and punctuation stripped) before lookup.
COLUMN_MAP: dict[str, str] = {
    "ser": "serial",
    "rollno": "roll_number",
    "name": "name",
    "fathername": "father_name",
    "authorizationno": "authorization_no",
    "email": "email",
    "testtype": "test_type",
    "contactno": "contact_no",
    "gender": "gender",
    "city": "city",
    "examcity": "exam_city",
    "graudatingcountry": "graduating_country",  # sic: spelling from the sheet
    "graduatingcountry": "graduating_country",
    "graduatingcollegename": "graduating_college",
    "sessionmorningevening": "session",
    "reportingtime": "reporting_time",
    "testtiming": "test_timing",
    "examcenter": "exam_center",
    "centerlocation": "center_location",
    "testmonth": "test_month",
}

AVATAR_SIZE = 128
AVATAR_PALETTE = [
    (31, 94, 255),
    (18, 183, 106),
    (217, 45, 32),
    (127, 86, 217),
    (247, 144, 9),
]


def normalise_header(value: Any) -> str:
    return "".join(ch for ch in str(value or "").lower() if ch.isalnum())


def clean(value: Any) -> str:
    """Excel cell -> trimmed string, with dates and floats made readable."""
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.strftime("%B %Y")
    if isinstance(value, dt.date):
        return value.strftime("%B %Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def initials_of(name: str) -> str:
    parts = [p for p in name.split() if p]
    if not parts:
        return "?"
    if len(parts) == 1:
        return parts[0][:2].upper()
    return (parts[0][0] + parts[-1][0]).upper()


def build_sample_avatar(name: str, seed: int) -> str:
    """Base64 PNG placeholder photo: initials on a flat colour disc."""
    colour = AVATAR_PALETTE[seed % len(AVATAR_PALETTE)]
    image = Image.new("RGB", (AVATAR_SIZE, AVATAR_SIZE), colour)
    draw = ImageDraw.Draw(image)
    draw.ellipse(
        (8, 8, AVATAR_SIZE - 8, AVATAR_SIZE - 8),
        fill=tuple(min(255, channel + 28) for channel in colour),
    )

    text = initials_of(name)
    try:
        font = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial.ttf", 52)
    except OSError:
        font = ImageFont.load_default()

    left, top, right, bottom = draw.textbbox((0, 0), text, font=font)
    draw.text(
        ((AVATAR_SIZE - (right - left)) / 2 - left,
         (AVATAR_SIZE - (bottom - top)) / 2 - top),
        text,
        font=font,
        fill=(255, 255, 255),
    )

    buffer = io.BytesIO()
    image.save(buffer, format="PNG", optimize=True)
    return base64.b64encode(buffer.getvalue()).decode("ascii")


class Command(BaseCommand):
    help = "Import candidates from an exam-centre .xlsx file."

    def add_arguments(self, parser) -> None:
        parser.add_argument(
            "path",
            nargs="?",
            default="data/islamabad.xlsx",
            help="Path to the .xlsx file (default: data/islamabad.xlsx)",
        )
        parser.add_argument(
            "--sheet",
            default=None,
            help="Worksheet name (default: the active sheet)",
        )

    def handle(self, *args: Any, **options: Any) -> None:
        path = Path(options["path"])
        if not path.exists():
            raise CommandError(f"Spreadsheet not found: {path}")

        workbook = load_workbook(path, data_only=True)
        sheet = workbook[options["sheet"]] if options["sheet"] else workbook.active

        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            raise CommandError("The spreadsheet is empty.")

        fields_by_index: dict[int, str] = {}
        unknown: list[str] = []
        for index, header in enumerate(header_row):
            key = normalise_header(header)
            if not key:
                continue
            field = COLUMN_MAP.get(key)
            if field:
                fields_by_index[index] = field
            else:
                unknown.append(str(header))

        if "roll_number" not in fields_by_index.values():
            raise CommandError("No 'Roll No' column found in the spreadsheet.")
        if unknown:
            self.stdout.write(self.style.WARNING(f"Ignored columns: {unknown}"))

        created = updated = skipped = 0

        with transaction.atomic():
            for row_number, row in enumerate(rows, start=2):
                values = {
                    field: clean(row[index])
                    for index, field in fields_by_index.items()
                    if index < len(row)
                }

                roll_number = values.pop("roll_number", "")
                if not roll_number:
                    skipped += 1
                    continue

                serial = values.pop("serial", "") or str(row_number - 1)
                values["serial"] = int(float(serial)) if serial else row_number - 1
                values["image"] = build_sample_avatar(
                    values.get("name", ""), values["serial"]
                )

                _, was_created = Candidate.objects.update_or_create(
                    roll_number=roll_number,
                    defaults=values,
                )
                created += was_created
                updated += not was_created

        self.stdout.write(
            self.style.SUCCESS(
                f"Imported {created} new, updated {updated}, skipped {skipped} rows "
                f"from {path}"
            )
        )
